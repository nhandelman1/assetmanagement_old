""" openpyxl utility functions

"""
import openpyxl.styles as OPStyles
import openpyxl.utils.cell as OPCellUtil
import os
import pandas as pd
from collections import defaultdict
from enum import Enum
from openpyxl import load_workbook
from typing import Optional, Union


# 0.5 for each comma and the decimal point
# 1   for "-" neg format
# 0.5 for "(" and ")" neg format
# 1   for whitespace
# 1.1 for all other characters
WIDTH_DICT = defaultdict(lambda: 1.1, {"-": 1, "(": 0.5, ")": 0.5, ".": 0.5, ",": 0.5, " ": 1})


def calc_cell_width(value=None, left_count: int = 0, right_count=None, comma_fmt=False, neg_fmt=None):
    """ Determine cell width based on arguments

    Args:
        value (Optional[str, float, int]): calc width based on value provided. Default None
            If value is a str, no other arguments are considered
            If value is an int, comma_fmt and neg_fmt are considered.
            If value is a float, right_count, comma_fmt and neg_fmt are considered.
            If value is None, all other arguments are considered.
        left_count (int): count of digits to the left of the decimal point. >= 0. Default 0.
        right_count (Optional[int]): count of digits to the right of the decimal point. >= 0. Default None.
            If value is None and right_count is None, right_count defaults to 0.
            If value is str or int, right_count is ignored.
            If value is float and right_count is not None, the value of right_count overrides the number of digits
                in the fractional part of value.
        comma_fmt (Optional[boolean]): consider commas when determining number width. Default None for False
        neg_fmt (Optional[str]): negative format to consider for numbers. "-", "()", None. Default None
    """
    if comma_fmt is None:
        comma_fmt = False

    width = 0
    use_fmts = True

    if value is None:
        right_count = 0 if right_count is None else right_count
        value = "0" * left_count + ("" if right_count == 0 else ".") + "0" * right_count
    else:
        if isinstance(value, int):
            value = str(abs(value))
        elif isinstance(value, float):
            value = str(abs(value))

            if right_count is not None:
                try:
                    # value in scientific notation (e.g. 5e-05) breaks this
                    value = value[0: value.index(".")] + ("" if right_count == 0 else ".") + "0" * right_count
                except ValueError:
                    # TODO make this work correctly with scientific notation
                    value = "0" + ("" if right_count == 0 else ".") + "0" * right_count
        else:  # value is a str
            use_fmts = False

    if use_fmts:
        if neg_fmt is not None:
            value += neg_fmt
        if comma_fmt:
            left_count = max(len(value.split(".")[0]), 1)
            value += ("," * ((left_count - 1) // 3))

    for c in value:
        width += WIDTH_DICT[c]

    return width + 0.5


def sheet_adj_col_width(sheet, start_row: int = 1, end_row=None, min_width: float = 5, max_width: float = 50,
                        right_count=None, comma_fmt=None, neg_fmt=None):
    """ Auto adjust all column widths based on arguments

    Args:
        sheet (openpyxl.worksheet.worksheet.worksheet):
        start_row (Union[int, dict, defaultdict]): Default 1
            int for all columns to start width calculations at this row
            dict to start width calculation for the specified column (key) at specified row (value). Columns not
                included as keys will start at row 1. e.g. {"A": 1, "C": 2}
            defaultdict for dict with caller specified row for missing keys
        end_row (Optional[int, dict]): Default None
            None for all columns to end width calculations at the last row in each column
            int for all columns to end width calculations at this row
            dict to end width calculation for the specified column (key) at specified row (value). For columns not
                included as keys, calculation ends at last row in each of those columns. e.g. {"A": 1, "C": 2}
            defaultdict for dict with caller specified row for missing keys
        min_width (Union[float, defaultdict]): Default 5
            defaultdict to set by column (key) the min width (value) with caller specified min width for missing keys
        max_width (Union[float, defaultdict]): Default 50
            defaultdict to set by column (key) the max width (value) with caller specified max width for missing keys
        right_count (Optional[int, defaultdict]): For numeric columns, consider this many fractional digits.
            For default None, use the fractional part of the value in each cell in the column
            defaultdict to set by column (key) the count (value) with caller specified count for missing keys
        comma_fmt (Optional[boolean, defaultdict]): For numeric columns, consider commas. Default None for False.
            defaultdict to set by column (key) the boolean (value) with caller specified boolean for missing keys
        neg_fmt (Optional[str, defaultdict]): For numeric columns, consider "-", "()", None fmts. Default None.
            defaultdict to set by column (key) the format (value) with caller specified format for missing keys
    """
    if comma_fmt is None:
        comma_fmt = False

    def help_1(arg, def_val=None):
        b = isinstance(arg, dict)
        if b:
            val = None
            if not isinstance(arg, defaultdict):
                arg = defaultdict(lambda: def_val, arg)
        else:
            val = arg
        return b, arg, val

    def help_2(arg, ret_val=None):
        b = isinstance(arg, defaultdict)
        return b, ret_val if b else arg

    start_row_is_dict, start_row, s_r = help_1(start_row, def_val=1)
    end_row_is_dict, end_row, e_r = help_1(end_row)

    min_width_is_dict, min_w_for_col = help_2(min_width)
    max_width_is_dict, max_w_for_col = help_2(max_width)
    right_count_is_dict, right_count_for_col = help_2(right_count)
    comma_fmt_is_dict, comma_fmt_for_col = help_2(comma_fmt, ret_val=False)
    neg_fmt_is_dict, neg_fmt_for_col = help_2(neg_fmt)

    for col_index in range(1, sheet.max_column+1):
        col_letter = OPCellUtil.get_column_letter(col_index)
        s_r = start_row[col_letter] if start_row_is_dict else s_r
        right_count_for_col = right_count[col_letter] if right_count_is_dict else right_count_for_col
        comma_fmt_for_col = comma_fmt[col_letter] if comma_fmt_is_dict else comma_fmt_for_col
        neg_fmt_for_col = neg_fmt[col_letter] if neg_fmt_is_dict else neg_fmt_for_col
        min_w_for_col = min_width[col_letter] if min_width_is_dict else min_w_for_col
        max_w_for_col = max_width[col_letter] if max_width_is_dict else max_w_for_col

        if end_row_is_dict:
            e_r = end_row[col_letter]
        if e_r is None or end_row is None:
            e_r = len(sheet[col_letter])

        col_w = min_w_for_col
        for cell_tuple in sheet[col_letter + str(s_r): col_letter + str(e_r)]:
            cell = cell_tuple[0]
            if cell.data_type in ("s", "d"):
                w = calc_cell_width(value=str(cell.value))
            else:  # cell.data_type == "n"
                w = calc_cell_width(value=cell.value, right_count=right_count_for_col, comma_fmt=comma_fmt_for_col,
                                    neg_fmt=neg_fmt_for_col)
            col_w = max(col_w, w)

        col_w = min(col_w, max_w_for_col)

        sheet.column_dimensions[col_letter].width = col_w


def df_to_file_sheet(df, file_name, sheet_name, ret_writer: bool = False):
    """ Write pandas dataframe to excel file and sheet

    Args:
        df (pd.DataFrame): df to write to file
        file_name (str): name of output file
        sheet_name (str): name of output sheet in file
        ret_writer (boolean): True to leave writer open and return instance. False to close writer and return None.
            Default False

    Returns:
        Optional[pd.ExcelWriter]: dependent on value of ret_writer
    """
    if os.path.exists(file_name):
        # keep lines in this order
        book = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name, engine="openpyxl")
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
    else:
        writer = pd.ExcelWriter(file_name, engine="openpyxl")

    df.to_excel(writer, sheet_name=sheet_name)

    if ret_writer:
        return writer
    else:
        writer.close()


def clean_file_name(file_name, rep: str = " "):
    """ Replace disallowed excel file name characters with rep

    Args:
        file_name (str): name of file
        rep (str): replace disallowed characters with this value. Default " "

    Returns:
        str: file_name with disallowed characters replaced with rep
    """
    lst = ["\\", "/", ":", "*", "?", "\"", "<", ">", "|"]
    rd = defaultdict(lambda: rep)

    for c in lst:
        file_name = file_name.replace(c, rd[c])

    return file_name


class NumFmt(Enum):
    """ Enum of common formats

    These are Excel custom format strings. Also, see openpyxl.styles.numbers.
    """
    # Excel custom format. see openpyxl.styles.numbers
    GEN = "General"
    USD_COMMA_0 = "$#,##0"
    COMMA_2 = "#,##0.00"
    COMMA_4 = "#,##0.0000"
    ROUND_4 = "#0.0000"
    ROUND_6 = "#0.000000"
    SD_DASH = "yyyy-mm-dd"
    SD_SLASH = "yyyy/mm/dd"


class Horiz(Enum):
    LEFT = "left"
    CENT = "center"
    RIGHT = "right"
    JUST = "justify"
    NONE = None


class Vert(Enum):
    TOP = "top"
    CENT = "center"
    BOTT = "bottom"
    JUST = "justify"
    NONE = None


class Color(Enum):
    BLACK = "00000000"
    WHITE = "00FFFFFF"
    RED = "00FF0000"
    BLUE = "000000FF"
    GREEN = "0000FF00"
    YELLOW = "00FFFF00"
    ORANGE = "FFA500"


class CellStyle:
    """ Apply styles to specified cells

    See openpyxl.styles, openpyxl.styles.alignment, openpyxl.styles.borders, openpyxl.styles.fills,
    openpyxl.styles.fonts, openpyxl.styles.numbers
    Columns can be referenced by letter (e.g. "A", "AA") or by index (starting at 1)

    Attributes:
        See __init__ docstring
        cell_list (list[str]): e.g. ["A1", "C5"]
    """
    # TODO implement border, font
    # TODO apply styles without overriding existing styles
    def __init__(self, sheet, alignment=None, border=None, fill=None, font=None, num_fmt=None, horiz=None, vert=None,
                 fill_color=None):
        """ __init__ function

        Any optional arguments left as None will not be applied.

        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet):
            alignment (Optional[openpyxl.styles.alignment]): text alignment in cell
            border (Optional[openpyxl.styles.borders]): not implemented
            fill (Optional[openpyxl.styles.fills, openpyxl.styles.GradientFill]):
                The 'Color' options are for the background color only.
            font (Optional[openpyxl.styles.fonts]): not implemented
            num_fmt (Optional[NumFmt]): number format. Default None for "General" format
            horiz (Optional[Horiz]): overrides alignment horizontal if alignment is not None
            vert (Optional[Vert]): overrides alignment vertical if alignment is not None
            fill_color (Optional[openpyxl.styles.Color, Color]): overrides fill start_color and end_color if fill is
                not None
        """
        self.sheet = sheet

        if alignment is None and (horiz is not None or vert is not None):
            alignment = OPStyles.Alignment()
        if horiz is not None:
            alignment.horizontal = horiz.value
        if vert is not None:
            alignment.vertical = vert.value
        self.alignment = alignment

        self.border = None  # OPStyles.Border()

        if fill is None and fill_color is not None:
            fill = OPStyles.PatternFill()
        if fill_color is not None:
            if isinstance(fill_color, Color):
                fill_color = OPStyles.Color(rgb=fill_color.value)
            # must set fill_type and both fgColor and bgColor (these both default to black)
            fill.fill_type = "solid"
            fill.fgColor = fill_color
            fill.bgColor = fill_color
        self.fill = fill

        self.font = None  # OPStyles.Font()

        self.num_fmt = None if num_fmt is None else num_fmt.value

        self.cell_list = []

    def rows(self, rows, start_col: str = "A", end_col=None):
        """ Add specified cells to cell list

        Args:
            rows (Union[int, list[int]])
            start_col (Union[str, int]): Default "A".
            end_col (Optional[str, int]): None for all columns in each row. Default None.

        Returns:
            CellStyle: self for chaining function calls
        """
        if isinstance(rows, int):
            rows = [rows]

        if isinstance(start_col, str):
            start_col = OPCellUtil.column_index_from_string(start_col)

        if isinstance(end_col, str):
            end_col = OPCellUtil.column_index_from_string(end_col)

        for row in rows:
            end_c = len(self.sheet[row]) if end_col is None else end_col
            self.cell_list += [OPCellUtil.get_column_letter(c) + str(row) for c in range(start_col, end_c + 1)]

        return self

    def cols(self, cols, start_row: int = 1, end_row=None):
        """ Add specified cells to cell list

        Args:
            cols (Union[str, int, list[str], list[int]]):
            start_row (int): Default 1.
            end_row (Optional[int]): None for all rows in column. Default None.

        Returns:
            CellStyle: self for chaining function calls
        """
        if isinstance(cols, int):
            cols = [OPCellUtil.get_column_letter(cols)]
        elif isinstance(cols, list) and isinstance(cols[0], int):
            cols = [OPCellUtil.get_column_letter(c) for c in cols]

        for col in cols:
            end_r = len(self.sheet[col]) if end_row is None else end_row
            self.cell_list += [col + str(r) for r in range(start_row, end_r + 1)]

        return self

    def cells(self, cells):
        """ Add specified cells to cell list

        Args:
            cells (Union[str, list[str], list[int, int], list[list[int, int]]]): str format "A1" or list(int, int)
                format [row#, col#] or list(list) format [[row #, col#],...]

        Returns:
            CellStyle: self for chaining function calls
        """
        if isinstance(cells, str):
            self.cell_list.append(cells)
        else:  # list
            if isinstance(cells[0], int):  # [row#, col#]
                self.cell_list.append(OPCellUtil.get_column_letter(cells[1]) + str(cells[0]))
            elif isinstance(cells[0], list):  # [[row#, col#], ...]
                self.cell_list += [OPCellUtil.get_column_letter(rc[1]) + str(rc[0]) for rc in cells]
            else:  # ["A1", "C5", ...]
                self.cell_list += cells

        return self

    def apply(self):
        """ Apply styling to all cells in self.cell_list """
        for coord in self.cell_list:
            cell = self.sheet[coord]

            if self.alignment is not None:
                cell.alignment = self.alignment
            # cell.border = self.border
            if self.fill is not None:
                cell.fill = self.fill
            # cell.font = self.font
            if self.num_fmt is not None:
                cell.number_format = self.num_fmt

    def auto_num_fmt(self, ignore_excluded=True, all_cells=False):
        # auto fmt all cells in sheet except those in self.excluded_cells
        # elements in self.excluded_cells will have format "A1" or [row #, col #]

        # TODO implement this function
        pass


class CellFmt:
    """ Class for setting cell formats

    ********************************************************************************************************************
    DEPRECATED - USE CellStyle
    ********************************************************************************************************************

    """

    class Fmt(Enum):
        """ Enum of common formats

        These are Excel custom format strings. Also, see openpyxl.styles.numbers.
        """
        # Excel custom format. see openpyxl.styles.numbers
        GEN = "General"
        COMMA_2 = "#,##0.00"
        ROUND_4 = "#0.0000"
        SD_DASH = "yyyy-mm-dd"
        SD_SLASH = "yyyy/mm/dd"

    def __init__(self, sheet):
        """
        ****************************************************************************************************************
        DEPRECATED - USE CellStyle
        ****************************************************************************************************************

        """
        self.sheet = sheet
        self.excluded_cells = []

    def row_fmt(self, fmt, rows, start_col="A", end_col=None):
        """ Format by row

        Args:
            fmt (CellFmt.Fmt)
            rows (int, list(int))
            start_col (str, int, optional): Default "A".
            end_col (str, int, None, optional): None for all columns in row. Default None.
        """
        if isinstance(rows, int):
            rows = [rows]
        if isinstance(start_col, str):
            start_col = OPCellUtil.column_index_from_string(start_col)
        if isinstance(end_col, str):
            end_col = OPCellUtil.column_index_from_string(end_col)

        for row in rows:
            end_c = len(self.sheet[row]) if end_col is None else end_col

            for c in range(start_col, end_c + 1):
                # sheet.cells is expecting column to be a number (not a letter)
                self.sheet.cell(row, c).number_format = fmt.value
                self.excluded_cells.append([row, c])

    def col_fmt(self, fmt, cols, start_row=0, end_row=None):
        """ Format by column

        Args:
            fmt (CellFmt.Fmt)
            cols (str, int, list(str), list(int))
            start_row (int, optional): Default 0.
            end_row (int, None, optional): None for all rows in column. Default None.
        """
        if isinstance(cols, (int, str)):
            cols = [cols]

        for col in cols:
            if isinstance(col, str):
                col = OPCellUtil.column_index_from_string(col)

            end_r = len(self.sheet[OPCellUtil.get_column_letter(col)]) if end_row is None else end_row

            for r in range(start_row, end_r + 1):
                # sheet.cells is expecting column to be a number (not a letter)
                self.sheet.cell(r, col).number_format = fmt.value
                self.excluded_cells.append([r, col])

    def cell_fmt(self, fmt, cells):
        """ Format by cell

        Args:
            fmt (CellFmt.Fmt)
            cells (str, list(str), list(list)): str format "A1" or list(list) format [[row #, col#], [row #, col #]]
        """

        if isinstance(cells, str):
            cells = [cells]

        for cell in cells:
            # cell will have format "A1" or [row #, col #]
            if isinstance(cell, list):
                cell = OPCellUtil.get_column_letter(cell[1]) + str(cell[0])

            self.sheet[cell].number_format = fmt.value
            self.excluded_cells.append(cell)

    def auto_fmt(self, ignore_excluded=True):
        # elements in self.excluded_cells will have format "A1" or [row #, col #]
        pass